"""Build the final Sad section: audit-keeps (29) + Soaklist (31, deduped) = 58 unique.

Reads:
  - C:/Users/Branden/Downloads/sad_review 1.xlsx  (user-reviewed audit, only Y rows used)
  - C:/Users/Branden/Downloads/Soaklist.txt       (reference playlist for sulk DNA)

Writes:
  - inputs/sad_section_final.md       (markdown lines for taste_profile.md)
  - inputs/Soaklist.txt               (preserve in repo)
Then patches both desktop and repo taste_profile.md.
"""
from __future__ import annotations
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(r"C:\Users\Branden\OneDrive\Documents\Music Enrichment\Music-Enrichment")
DESKTOP_TASTE = Path(r"C:\Users\Branden\OneDrive\Desktop\taste_profile.md")
REPO_TASTE = REPO / "taste_profile.md"
REVIEW_XLSX = Path(r"C:\Users\Branden\Downloads\sad_review 1.xlsx")
SOAKLIST_SRC = Path(r"C:\Users\Branden\Downloads\Soaklist.txt")
SOAKLIST_DEST = REPO / "inputs" / "Soaklist.txt"
OUT_MD = REPO / "inputs" / "sad_section_final.md"


def tight(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def main() -> None:
    # 1. Pull keeps from the user-reviewed xlsx
    wb = load_workbook(REVIEW_XLSX, data_only=True)
    ws = wb["Sad review"]
    keeps: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] is None:
            continue
        keep = (str(row[0]).strip() if row[0] else "").upper()
        if keep in ("Y", "YES", "KEEP"):
            keeps.append((row[1], row[2]))
    print(f"Audit keeps: {len(keeps)}")

    # 2. Pull Soaklist tracks
    soak: list[tuple[str, str]] = []
    for line in SOAKLIST_SRC.read_text(encoding="utf-8-sig").splitlines():
        line = line.strip()
        if not line or " - " not in line:
            continue
        a, t = (p.strip() for p in line.split(" - ", 1))
        soak.append((a, t))
    print(f"Soaklist: {len(soak)}")

    # 3. Merge with dedup
    seen: set[str] = set()
    merged: list[tuple[str, str]] = []
    for a, t in keeps + soak:
        k = f"{tight(a)}|{tight(t)}"
        if k in seen:
            continue
        seen.add(k)
        merged.append((a, t))
    print(f"Merged unique: {len(merged)}")

    # 4. Sort alphabetical by artist
    merged.sort(key=lambda r: (r[0].lower(), r[1].lower()))

    # 5. Write final markdown lines
    lines = [f"- \"{t}\" by {a}" for a, t in merged]
    OUT_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUT_MD}")

    # 6. Preserve Soaklist in repo
    SOAKLIST_DEST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SOAKLIST_SRC, SOAKLIST_DEST)
    print(f"Copied Soaklist -> {SOAKLIST_DEST}")

    # 7. Patch desktop + repo taste_profile.md
    for target in (DESKTOP_TASTE, REPO_TASTE):
        text_lines = target.read_text(encoding="utf-8").splitlines()
        start_idx = None
        for i, line in enumerate(text_lines):
            if line.strip() == "### sad (locked)":
                start_idx = i
                break
        if start_idx is None:
            raise SystemExit(f"can't find ### sad (locked) in {target}")
        # Find next ### heading
        end_idx = None
        for i in range(start_idx + 1, len(text_lines)):
            if text_lines[i].startswith("### "):
                end_idx = i
                break
        if end_idx is None:
            raise SystemExit(f"can't find next section after sad in {target}")
        new_block = ["### sad (locked)"] + lines + [""]
        new_text = "\n".join(text_lines[:start_idx] + new_block + text_lines[end_idx:]) + "\n"
        target.write_text(new_text, encoding="utf-8")
        print(f"Patched {target} ({target.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
