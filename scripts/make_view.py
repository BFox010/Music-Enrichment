"""Generate a spreadsheet view of tracks.jsonl for analyst inspection.

Output goes to ``views/`` (gitignored). JSONL remains the source of truth.
Excel format provides filterable columns, frozen headers, and column auto-fit.

Usage:
    python scripts/make_view.py                           # full library
    python scripts/make_view.py --top 200                 # top-200 by play_count
    python scripts/make_view.py --output views/recent.xlsx --since 2025-01-01
    python scripts/make_view.py --format csv              # CSV instead of XLSX
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path

# Ensure repo root is importable when run as `python scripts/make_view.py`
REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from pipeline.config import TRACKS_PATH, VIEWS_DIR  # noqa: E402

# Columns shown in the spreadsheet, in order.
# Lists are flattened to comma-joined strings; nested dicts are expanded.
COLUMNS: list[str] = [
    "artist", "track", "album", "play_count", "peak_year", "release_year",
    "first_scrobbled", "last_scrobbled",
    "lastfm_tags", "itunes_genre", "genres",
    "duration_ms", "explicit",
    "musicbrainz_id", "spotify_id", "apple_music_available", "apple_music_id",
    "audio_features.danceability", "audio_features.energy",
    "audio_features.valence", "audio_features.tempo",
    "audio_features.loudness",
    "mood_tags", "mood_source", "mood_confidence",
    "itunes_play_count", "itunes_skip_count", "itunes_date_added",
    "saturation_tier", "blacklisted", "playlists",
    "curation_state", "rejected_reason",
    "enriched_at",
]


def _flatten(row: dict, key: str) -> object:
    """Get a possibly-nested value (dot-notation) from a track row."""
    if "." not in key:
        value = row.get(key)
    else:
        head, _, tail = key.partition(".")
        sub = row.get(head)
        if isinstance(sub, dict):
            value = sub.get(tail)
        else:
            value = None
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if value is None:
        return ""
    return value


def _load_tracks(path: Path) -> list[dict]:
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _filter_rows(
    rows: list[dict],
    *,
    top: int | None,
    since: str | None,
    artist_contains: str | None,
) -> list[dict]:
    out = list(rows)
    if since:
        out = [r for r in out if (r.get("last_scrobbled") or "") >= since]
    if artist_contains:
        needle = artist_contains.lower()
        out = [r for r in out if needle in (r.get("artist") or "").lower()]
    out.sort(key=lambda r: -int(r.get("play_count") or 0))
    if top:
        out = out[:top]
    return out


def _write_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(COLUMNS)
        for row in rows:
            writer.writerow([_flatten(row, col) for col in COLUMNS])


def _write_xlsx(rows: list[dict], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl missing — install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "tracks"

    # Header row — bold, dark fill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F4F")
    for col_idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, col in enumerate(COLUMNS, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_flatten(row, col))

    # Freeze top row, auto-filter, column widths
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    for c_idx, col in enumerate(COLUMNS, start=1):
        # heuristic width based on column name length
        ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(40, len(col) + 4))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def make_view(
    *,
    tracks_path: Path = TRACKS_PATH,
    output: Path | None = None,
    fmt: str = "xlsx",
    top: int | None = None,
    since: str | None = None,
    artist_contains: str | None = None,
) -> Path:
    if not tracks_path.exists():
        raise FileNotFoundError(f"{tracks_path} not found — run the pipeline first.")

    rows = _load_tracks(tracks_path)
    rows = _filter_rows(rows, top=top, since=since, artist_contains=artist_contains)

    if output is None:
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        suffix = "csv" if fmt == "csv" else "xlsx"
        output = VIEWS_DIR / f"tracks_{ts}.{suffix}"

    if fmt == "csv":
        output.parent.mkdir(parents=True, exist_ok=True)
        _write_csv(rows, output)
    else:
        _write_xlsx(rows, output)

    print(f"Wrote {len(rows)} rows -> {output}")
    return output


def _parse_args(argv: list[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Generate a tracks.jsonl spreadsheet view.")
    p.add_argument("--tracks", type=Path, default=TRACKS_PATH)
    p.add_argument("--output", type=Path, default=None)
    p.add_argument("--format", choices=["xlsx", "csv"], default="xlsx")
    p.add_argument("--top", type=int, default=None,
                   help="Limit to top-N rows by play_count")
    p.add_argument("--since", default=None,
                   help="Only tracks last_scrobbled on or after this date (YYYY-MM-DD)")
    p.add_argument("--artist", dest="artist_contains", default=None,
                   help="Filter to artists whose name contains this substring")
    return p.parse_args(argv)


if __name__ == "__main__":
    args = _parse_args(sys.argv[1:])
    make_view(
        tracks_path=args.tracks,
        output=args.output,
        fmt=args.format,
        top=args.top,
        since=args.since,
        artist_contains=args.artist_contains,
    )
