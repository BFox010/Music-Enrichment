"""Build a spot-check xlsx for the 13 non-Sad mood categories.

For each category, samples 5 audit-source + 5 centroid-source tracks
(or all available if <5), sorted by play count desc. Centroid rows are
the suspicious ones — same hallucination pattern that produced bad Sad.

Output: inputs/mood_spotcheck.xlsx
  - "Spot check" sheet: ~130 rows, Keep? dropdown, color-coded by source
  - "Instructions" sheet: quick reference for the workflow
"""
from __future__ import annotations
import json
import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

REPO = Path(r"C:\Users\Branden\OneDrive\Documents\Music Enrichment\Music-Enrichment")
TRACKS = REPO / "tracks.jsonl"
OUT = REPO / "inputs" / "mood_spotcheck.xlsx"

# 13 categories to spot-check (Sad excluded — already curated)
CATEGORIES = [
    "Uplifting", "Sunny", "Happy", "Love",
    "Hype", "Fast", "Dance", "Groove",
    "Heavy Bass", "Moody", "Slow", "Heartbreak", "Dark",
]

SAMPLES_PER_SOURCE = 5
random.seed(42)  # reproducible


def main() -> None:
    tracks = [json.loads(line) for line in TRACKS.read_text(encoding="utf-8").splitlines() if line.strip()]

    rows: list[dict] = []
    for cat in CATEGORIES:
        in_cat = [t for t in tracks if cat in (t.get("mood_tags") or [])]
        audit = sorted(
            [t for t in in_cat if t.get("mood_source") == "audit"],
            key=lambda t: -(t.get("play_count") or 0),
        )
        centroid = sorted(
            [t for t in in_cat if t.get("mood_source") == "centroid"],
            key=lambda t: -(t.get("play_count") or 0),
        )
        # Take top-played from each source so user sees familiar tracks
        picks = audit[:SAMPLES_PER_SOURCE] + centroid[:SAMPLES_PER_SOURCE]
        for t in picks:
            other = [m for m in (t.get("mood_tags") or []) if m != cat]
            rows.append({
                "category": cat,
                "artist": t["artist"],
                "track": t["track"],
                "source": t.get("mood_source") or "?",
                "other_moods": ", ".join(other),
                "plays": t.get("play_count") or 0,
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Spot check"

    headers = ["Keep?", "Category", "Artist", "Track", "Source", "Other moods", "Plays", "Notes"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", start_color="3F3F3F")
        c.alignment = Alignment(horizontal="center", vertical="center")

    src_fills = {
        "audit": PatternFill("solid", start_color="D5E8D4"),    # green = trusted
        "centroid": PatternFill("solid", start_color="FFE6CC"),  # orange = suspect
    }
    cat_fills = {}
    palette = ["FFF2CC", "DAE8FC", "F8CECC", "D4E1F5", "E1D5E7", "FFF4E6", "D5E8D4"]
    for i, cat in enumerate(CATEGORIES):
        cat_fills[cat] = PatternFill("solid", start_color=palette[i % len(palette)])

    for r in rows:
        ws.append(["", r["category"], r["artist"], r["track"], r["source"], r["other_moods"], r["plays"], ""])
        row_idx = ws.max_row
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(vertical="center", wrap_text=False)
        # Color the source cell
        if r["source"] in src_fills:
            ws.cell(row=row_idx, column=5).fill = src_fills[r["source"]]
        # Tint category cell
        ws.cell(row=row_idx, column=2).fill = cat_fills[r["category"]]

    # Y/N dropdown on Keep?
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv.error = "Type Y or N"
    dv.errorTitle = "Invalid"
    ws.add_data_validation(dv)
    dv.add(f"A2:A{ws.max_row}")

    widths = {"A": 8, "B": 12, "C": 24, "D": 42, "E": 10, "F": 26, "G": 7, "H": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Set row height for readability on phones
    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "C2"  # freeze first 2 cols + header
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    # Instructions sheet
    instr = wb.create_sheet("Instructions")
    instr["A1"] = "Mood spot-check — quick workflow"
    instr["A1"].font = Font(name="Arial", bold=True, size=14)
    notes = [
        "",
        "Goal: for each of 13 mood categories, judge whether the track really",
        "belongs to that mood. Same question we asked for Sad.",
        "",
        "For each row, type Y or N in column A (or leave blank to skip).",
        "  Y = yes, this track legitimately fits the Category",
        "  N = no, this is wrong (centroid hallucination or audit co-tag bleed)",
        "",
        "Source column color guide:",
        "  green  = audit  — YOUR explicit label. High confidence.",
        "  orange = centroid — ML prediction. The suspicious ones.",
        "",
        "Sample size per category: 5 audit + 5 centroid (top-played from each),",
        "for ~130 rows total. Doable in a sitting.",
        "",
        "Verdict heuristic afterward:",
        "  If a category has 8-10 Y's overall (especially among centroid rows),",
        "    it's clean — leave alone.",
        "  If a category has many N's in the centroid block but Y's in audit,",
        "    Phase 6 centroid is hallucinating for that category — needs",
        "    Sad-style re-curation: drop centroid predictions, source only",
        "    from your audit.",
        "  If even the audit rows have N's, the audit itself drifted for that",
        "    category and may need re-tagging.",
        "",
        "When done: save the file, send back to Claude. Claude will summarize",
        "per-category and propose next moves.",
        "",
        f"Sampling seed: 42 (deterministic — re-running the script gives same rows)",
    ]
    for i, n in enumerate(notes, start=2):
        instr.cell(row=i, column=1).value = n
        instr.cell(row=i, column=1).font = Font(name="Arial")
    instr.column_dimensions["A"].width = 80

    wb.save(OUT)
    print(f"Wrote {len(rows)} rows across {len(CATEGORIES)} categories to {OUT}")
    by_cat = {}
    for r in rows:
        by_cat.setdefault(r["category"], {"audit": 0, "centroid": 0})
        by_cat[r["category"]][r["source"]] = by_cat[r["category"]].get(r["source"], 0) + 1
    for cat, c in by_cat.items():
        print(f"  {cat}: audit={c.get('audit', 0)} centroid={c.get('centroid', 0)}")


if __name__ == "__main__":
    main()
